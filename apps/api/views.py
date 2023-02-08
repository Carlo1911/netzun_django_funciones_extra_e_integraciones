from base64 import b64encode
from io import BytesIO

import qrcode
import qrcode.image.svg
import requests
import xlsxwriter
from django.core.mail import send_mail
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from rest_framework.views import APIView
from weasyprint import CSS, HTML
from weasyprint.text.fonts import FontConfiguration


class MailtrapEmail(APIView):
    def post(self, request):
        name = request.data.get("name")
        email = request.data.get("email")

        send_mail(
            "Asunto: Correo de prueba",
            f"Cuerpo: Hola {name}, este es un correo de prueba",
            "carlo.alva@yourdjangoapp.com",
            [email],
            fail_silently=False,
        )

        return JsonResponse({"result": "email sent"})


class MailtrapTemplateEmail(APIView):
    def post(self, request):
        data = request.data

        html_message = render_to_string("email_template.html", {"name": data["name"]})
        plain_message = strip_tags(html_message)
        send_mail(
            "Asunto: Correo de prueba con template",
            plain_message,
            "carlo.alva@yourdjangoapp.com",
            [data["email"]],
            html_message=html_message,
        )

        return JsonResponse({"result": "email sent"})


class HotmailProviderEmail(APIView):
    def post(self, request):
        name = request.data.get("name")
        email = request.data.get("email")

        send_mail(
            "Asunto: Correo de prueba",
            f"Cuerpo: Hola {name}, este es un correo de prueba desde hotmail",
            "carlo0071@hotmail.com",
            [email],
            fail_silently=False,
        )

        return JsonResponse({"result": "email sent"})


class GmailProviderEmail(APIView):
    def post(self, request):
        name = request.data.get("name")
        email = request.data.get("email")

        send_mail(
            "Asunto: Correo de prueba",
            f"Cuerpo: Hola {name}, este es un correo de prueba desde gmail",
            "carlo0071@gmail.com",
            [email],
            fail_silently=False,
        )

        return JsonResponse({"result": "email sent"})


class MailgunAnymailProviderEmail(APIView):
    def post(self, request):
        name = request.data.get("name")
        email = request.data.get("email")

        send_mail(
            "Asunto: Correo de prueba",
            f"Cuerpo: Hola {name}, este es un correo de prueba desde mailgun usando anymail",
            "carlo0071@gmail.com",
            [email],
            fail_silently=False,
        )

        return JsonResponse({"result": "email sent"})


class MailgunRequestsProviderEmail(APIView):
    def post(self, request):
        name = request.data.get("name")
        email = request.data.get("email")

        requests.post(
            "https://api.mailgun.net/v3/sandbox0519811b781943ff8ef4b01402d6c8ff.mailgun.org/messages",
            auth=("api", "665a4e0201c6f78276554b5dc52dcaec-48d7d97c-9aa7144b"),
            data={
                "from": "Mailgun Sandbox <postmaster@sandbox0519811b781943ff8ef4b01402d6c8ff.mailgun.org>",
                "to": email,
                "subject": "Correo de prueba",
                "text": f"Hola {name}, este es un correo de prueba desde mailgun usando requests",
            },
        )

        return JsonResponse({"result": "email sent"})


class QRGenerator(APIView):
    def get(self, request):
        # query params
        version = request.query_params.get("version", 1)
        box_size = request.query_params.get("box_size", 10)
        border = request.query_params.get("border", 5)
        data = request.query_params.get("data", "Django avanzado")

        qr = qrcode.QRCode(
            version=version,
            box_size=box_size,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            image_factory=qrcode.image.svg.SvgImage,
            border=border,
        )

        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image()
        stream = BytesIO()
        img.save(stream)

        img.save('qr.svg')

        return JsonResponse(
            {
                "result": "qr generated",
                "data": f"data:image/svg+xml;charset=utf-8;base64,{b64encode(img.to_string()).decode()}",
            }
        )


class OpenpyxlExcel(APIView):
    def post(self, request):
        students = request.data.get("students", "[]")

        wb = Workbook()
        ws = wb.active

        ws.merge_cells("A1:C1")
        ws["A1"] = "Django avanzado"
        ws["A1"].font = Font(bold=True, size=20)
        ws["A1"].alignment = Alignment(horizontal="center")
        # headers
        ws["A2"] = "Nombre"
        ws["B2"] = "Apellido"
        ws["C2"] = "Edad"

        for i, student in enumerate(students):
            for j, value in enumerate(student.values()):
                ws.cell(row=i + 3, column=j + 1).value = value

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="excel_con_openpyxl.xlsx"'})

        wb.save(response)

        return response


class XlsxwriterExcel(APIView):
    def post(self, request):
        students = request.data.get("students", "[]")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="excel_con_xlsxwriter.xlsx"'})

        wb = xlsxwriter.Workbook(response, {'in_memory': True})
        ws = wb.add_worksheet()

        format_title = wb.add_format({'bold': True, 'font_size': 20, 'align': 'center'})
        ws.merge_range("A1:C1", 'Django avanzado', format_title)
        # headers
        ws.write("A2","Nombre")
        ws.write("B2","Apellido")
        ws.write("C2","Edad")

        for i, student in enumerate(students):
            for j, value in enumerate(student.values()):
                ws.write(i + 2, j, value)

        wb.close()

        return response

class PDFGenerator(APIView):
    def post(self, request):
        students = request.data.get("students", "[]")

        font_config = FontConfiguration()

        css = CSS(
            string="""
        @page {
            size: A4;
            margin: 0cm;
            }
        """,
            font_config=font_config,
        )
        html = render_to_string("pdf_template.html", {"students": students})
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="pdf_weasyprint.pdf"'
        HTML(string=html).write_pdf(response, stylesheets=[css], font_config=font_config)
        return response
